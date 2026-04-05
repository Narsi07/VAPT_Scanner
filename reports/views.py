# -*- coding: utf-8 -*-
import csv
import io
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.views import View
from django.shortcuts import render

from tools.models import NmapResultDb, NmapScanDb, SslscanResultDb
from staticscanners.models import StaticScanResultsDb
from projects.models import ProjectDb


def _get_org(request):
    return request.user.organization


@method_decorator(login_required, name="dispatch")
class ReportIndex(View):
    """Report generation landing page."""

    def get(self, request):
        org = _get_org(request)
        projects = ProjectDb.objects.filter(organization=org)
        # Summary counts
        nmap_count = NmapScanDb.objects.filter(organization=org).count()
        static_count = StaticScanResultsDb.objects.filter(organization=org).count()
        ssl_count = SslscanResultDb.objects.filter(organization=org).count()
        return render(request, "reports/index.html", {
            "projects": projects,
            "nmap_count": nmap_count,
            "static_count": static_count,
            "ssl_count": ssl_count,
        })


@method_decorator(login_required, name="dispatch")
class ExportCSV(View):
    """Export all scan results as a CSV file."""

    def get(self, request):
        org = _get_org(request)
        project_uuid = request.GET.get("project_id")

        # Resolve UUID → integer pk (NmapScanDb.project is a FK to ProjectDb.id)
        project_pk = None
        if project_uuid and project_uuid != "all":
            try:
                project_pk = ProjectDb.objects.get(uu_id=project_uuid, organization=org).pk
            except ProjectDb.DoesNotExist:
                project_pk = None

        response = HttpResponse(content_type="text/csv")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response["Content-Disposition"] = f'attachment; filename="vapt_report_{timestamp}.csv"'

        writer = csv.writer(response)

        # ── Nmap Results ──────────────────────────────────────────────────────
        writer.writerow(["=== NMAP PORT SCAN RESULTS ==="])
        writer.writerow(["IP Address", "Port", "Protocol", "State", "Service", "Version", "CPE"])
        nmap_qs = NmapResultDb.objects.filter(organization=org)
        if project_pk is not None:
            scan_ids = NmapScanDb.objects.filter(
                organization=org, project_id=project_pk
            ).values_list("scan_id", flat=True)
            nmap_qs = nmap_qs.filter(scan_id__in=scan_ids)
        for r in nmap_qs:
            writer.writerow([
                r.ip_address, r.port, r.protocol, r.state,
                r.name, r.version, r.cpe
            ])

        writer.writerow([])

        # ── Static Analysis ───────────────────────────────────────────────────
        writer.writerow(["=== STATIC ANALYSIS RESULTS ==="])
        writer.writerow(["Title", "Severity", "File Path", "Scanner", "Status", "Date"])
        static_qs = StaticScanResultsDb.objects.filter(organization=org)
        if project_pk is not None:
            static_qs = static_qs.filter(project_id=project_pk)
        for r in static_qs:
            writer.writerow([
                r.title, r.severity, r.filePath or r.fileName,
                r.scanner, r.vuln_status, r.date_time
            ])

        writer.writerow([])

        # ── SSL Scan ──────────────────────────────────────────────────────────
        writer.writerow(["=== SSL SCAN RESULTS ==="])
        writer.writerow(["Scan URL", "Status", "Output Summary"])
        ssl_qs = SslscanResultDb.objects.filter(organization=org)
        if project_pk is not None:
            ssl_qs = ssl_qs.filter(project_id=project_pk)
        for r in ssl_qs:
            output_preview = (r.sslscan_output or "")[:200].replace("\n", " ")
            writer.writerow([r.scan_url, "Completed" if r.sslscan_output else "Pending", output_preview])

        return response


@method_decorator(login_required, name="dispatch")
class ExportExcel(View):
    """Export all scan results as an Excel (.xlsx) file."""

    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return HttpResponse(
                "[ERROR] openpyxl is not installed. Run: pip install openpyxl",
                status=500,
                content_type="text/plain"
            )

        org = _get_org(request)
        project_uuid = request.GET.get("project_id")

        # Resolve UUID → integer pk
        project_pk = None
        if project_uuid and project_uuid != "all":
            try:
                project_pk = ProjectDb.objects.get(uu_id=project_uuid, organization=org).pk
            except ProjectDb.DoesNotExist:
                project_pk = None

        wb = openpyxl.Workbook()

        # ── Styles ─────────────────────────────────────────────────────────────
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2563EB")
        center = Alignment(horizontal="center")

        def make_header(ws, headers):
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center

        # ── Sheet 1: Nmap ──────────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Nmap Port Scan"
        make_header(ws1, ["IP Address", "Port", "Protocol", "State", "Service", "Version", "CPE"])
        nmap_qs = NmapResultDb.objects.filter(organization=org)
        if project_pk is not None:
            scan_ids = NmapScanDb.objects.filter(
                organization=org, project_id=project_pk
            ).values_list("scan_id", flat=True)
            nmap_qs = nmap_qs.filter(scan_id__in=scan_ids)
        for r in nmap_qs:
            ws1.append([r.ip_address, r.port, r.protocol, r.state, r.name, r.version, r.cpe])

        # ── Sheet 2: Static Analysis ───────────────────────────────────────────
        ws2 = wb.create_sheet("Static Analysis")
        make_header(ws2, ["Title", "Severity", "File Path", "Scanner", "Status", "Date"])
        static_qs = StaticScanResultsDb.objects.filter(organization=org)
        if project_pk is not None:
            static_qs = static_qs.filter(project_id=project_pk)
        for r in static_qs:
            ws2.append([
                r.title, r.severity, r.filePath or r.fileName,
                r.scanner, r.vuln_status, str(r.date_time) if r.date_time else ""
            ])

        # ── Sheet 3: SSL Scan ──────────────────────────────────────────────────
        ws3 = wb.create_sheet("SSL Scan")
        make_header(ws3, ["Scan URL", "Status", "Output Preview"])
        ssl_qs = SslscanResultDb.objects.filter(organization=org)
        for r in ssl_qs:
            output_preview = (r.sslscan_output or "")[:300].replace("\n", " ")
            ws3.append([r.scan_url, "Completed" if r.sslscan_output else "Pending", output_preview])

        # ── Auto-fit columns ───────────────────────────────────────────────────
        for ws in [ws1, ws2, ws3]:
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        # ── Response ───────────────────────────────────────────────────────────
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="vapt_report_{timestamp}.xlsx"'
        wb.save(response)
        return response
